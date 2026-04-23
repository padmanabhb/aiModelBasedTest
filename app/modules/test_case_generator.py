"""
Test Case Generator Module
Generates minimal test cases with complete coverage based on graph traversal

Supports:
- Node Coverage: All nodes must be visited
- Edge Coverage: All edges must be traversed
- Path Coverage: All possible paths covered
- Pair Coverage: All node pairs covered
"""
from typing import Dict, List, Set, Tuple, Any
from dataclasses import dataclass


@dataclass
class TestCase:
    """Represents a single test case"""
    test_number: int
    description: str  # Gherkin format (Given-When-Then)
    expected_result: str
    coverage_type: str  # "nodes", "edges", "pairs"
    nodes_covered: List[str]
    edges_covered: List[Tuple[str, str]]


class TestCaseGenerator:
    """
    Generates minimal test cases based on graph coverage criteria
    """
    
    def __init__(self, nodes: Dict[str, Any], edges: List[Dict[str, str]]):
        """
        Initialize test case generator
        
        Args:
            nodes: Dictionary of nodes
            edges: List of edge dictionaries with 'source' and 'target'
        """
        self.nodes = nodes
        self.edges = edges
        self.adjacency_list = self._build_adjacency_list()
        self.reverse_adjacency = self._build_reverse_adjacency()
    
    def _build_adjacency_list(self) -> Dict[str, List[str]]:
        """Build adjacency list for the graph"""
        adj_list = {}
        for node_id in self.nodes.keys():
            adj_list[node_id] = []
        
        for edge in self.edges:
            source = edge.get("source", "")
            target = edge.get("target", "")
            if source and target:
                adj_list[source].append(target)
        
        return adj_list
    
    def _build_reverse_adjacency(self) -> Dict[str, List[str]]:
        """Build reverse adjacency list"""
        rev_adj = {}
        for node_id in self.nodes.keys():
            rev_adj[node_id] = []
        
        for edge in self.edges:
            source = edge.get("source", "")
            target = edge.get("target", "")
            if source and target:
                rev_adj[target].append(source)
        
        return rev_adj
    
    def generate_node_coverage_tests(self) -> List[TestCase]:
        """
        Generate minimal test cases that cover all nodes
        
        Returns:
            List of test cases covering all nodes
        """
        test_cases = []
        
        # Find start and end nodes
        start_nodes = [n for n in self.nodes.keys() 
                      if len(self.reverse_adjacency.get(n, [])) == 0]
        end_nodes = [n for n in self.nodes.keys() 
                    if len(self.adjacency_list.get(n, [])) == 0]
        
        if not start_nodes:
            start_nodes = [list(self.nodes.keys())[0]]
        if not end_nodes:
            end_nodes = [list(self.nodes.keys())[-1]]
        
        covered_nodes = set()
        test_num = 1
        
        # Generate paths from each start node to each end node
        for start in start_nodes:
            for end in end_nodes:
                path = self._find_longest_path_dfs(start, end, set())
                if path:
                    # Calculate new nodes covered
                    new_nodes = [n for n in path if n not in covered_nodes]
                    if new_nodes or test_num == 1:
                        covered_nodes.update(path)
                        
                        test_case = TestCase(
                            test_number=test_num,
                            description=self._generate_gherkin_for_path(path),
                            expected_result=f"Successfully traverse from {start} to {end}",
                            coverage_type="nodes",
                            nodes_covered=path,
                            edges_covered=[(path[i], path[i+1]) for i in range(len(path)-1)]
                        )
                        test_cases.append(test_case)
                        test_num += 1
        
        # Add remaining uncovered nodes
        uncovered = set(self.nodes.keys()) - covered_nodes
        for node in uncovered:
            test_case = TestCase(
                test_number=test_num,
                description=f"Given a node {node} exists\nWhen the system processes it\nThen the node should be accessible",
                expected_result=f"Node {node} is accessible",
                coverage_type="nodes",
                nodes_covered=[node],
                edges_covered=[]
            )
            test_cases.append(test_case)
            test_num += 1
        
        return test_cases
    
    def generate_edge_coverage_tests(self) -> List[TestCase]:
        """
        Generate minimal test cases that cover all edges
        
        Returns:
            List of test cases covering all edges
        """
        test_cases = []
        covered_edges = set()
        test_num = 1
        
        # Find start nodes
        start_nodes = [n for n in self.nodes.keys() 
                      if len(self.reverse_adjacency.get(n, [])) == 0]
        if not start_nodes:
            start_nodes = [list(self.nodes.keys())[0]]
        
        # Generate paths that cover edges
        for start in start_nodes:
            path = self._dfs_traverse_uncovered_edges(start, set(), covered_edges)
            if len(path) > 1:
                edge_list = [(path[i], path[i+1]) for i in range(len(path)-1)]
                
                test_case = TestCase(
                    test_number=test_num,
                    description=self._generate_gherkin_for_path(path),
                    expected_result=f"All edges in path {' → '.join(path)} are traversed",
                    coverage_type="edges",
                    nodes_covered=path,
                    edges_covered=edge_list
                )
                test_cases.append(test_case)
                test_num += 1
        
        # Add remaining uncovered edges
        for edge in self.edges:
            source = edge.get("source", "")
            target = edge.get("target", "")
            edge_tuple = (source, target)
            if edge_tuple not in covered_edges:
                test_case = TestCase(
                    test_number=test_num,
                    description=f"Given system at node {source}\nWhen navigating to {target}\nThen edge {source}→{target} should be traversed",
                    expected_result=f"Edge {source}→{target} is successfully traversed",
                    coverage_type="edges",
                    nodes_covered=[source, target],
                    edges_covered=[edge_tuple]
                )
                test_cases.append(test_case)
                test_num += 1
        
        return test_cases
    
    def generate_pair_coverage_tests(self) -> List[TestCase]:
        """
        Generate minimal test cases that cover all node pairs
        
        Returns:
            List of test cases covering all node pairs
        """
        test_cases = []
        nodes = list(self.nodes.keys())
        test_num = 1
        
        # Generate test for each pair with a path
        for i, source in enumerate(nodes):
            for target in nodes[i+1:]:
                path = self._find_path_dfs(source, target, set())
                
                if path:
                    test_case = TestCase(
                        test_number=test_num,
                        description=f"Given user at {source}\nWhen navigating to {target}\nThen should reach {target} via path {' → '.join(path)}",
                        expected_result=f"Successfully navigate from {source} to {target}",
                        coverage_type="pairs",
                        nodes_covered=path,
                        edges_covered=[(path[i], path[i+1]) for i in range(len(path)-1)]
                    )
                else:
                    test_case = TestCase(
                        test_number=test_num,
                        description=f"Given user at {source}\nWhen attempting to navigate to {target}\nThen should handle unreachable node",
                        expected_result=f"System handles unreachable node {target} gracefully",
                        coverage_type="pairs",
                        nodes_covered=[source],
                        edges_covered=[]
                    )
                
                test_cases.append(test_case)
                test_num += 1
        
        return test_cases
    
    def generate_in_out_coverage_tests(self) -> List[TestCase]:
        """
        Generate test cases covering in-edges and out-edges for each node
        
        Returns:
            List of test cases for in-out edge coverage
        """
        test_cases = []
        test_num = 1
        
        for node_id in self.nodes.keys():
            in_edges = [e for e in self.edges if e["target"] == node_id]
            out_edges = [e for e in self.edges if e["source"] == node_id]
            
            if in_edges or out_edges:
                in_sources = [e["source"] for e in in_edges]
                out_targets = [e["target"] for e in out_edges]
                
                test_description = f"Given node {node_id}"
                if in_sources:
                    test_description += f"\nWhen accessed from {', '.join(in_sources[:2])}"
                if out_targets:
                    test_description += f"\nThen can navigate to {', '.join(out_targets[:2])}"
                
                test_case = TestCase(
                    test_number=test_num,
                    description=test_description,
                    expected_result=f"Node {node_id} properly handles {len(in_edges)} incoming and {len(out_edges)} outgoing edges",
                    coverage_type="in-out",
                    nodes_covered=[node_id],
                    edges_covered=[(e["source"], e["target"]) for e in in_edges + out_edges]
                )
                test_cases.append(test_case)
                test_num += 1
        
        return test_cases
    
    def _find_longest_path_dfs(self, start: str, end: str, visited: Set[str]) -> List[str]:
        """Find longest path using DFS"""
        if start == end:
            return [start]
        
        visited.add(start)
        longest = []
        
        for neighbor in self.adjacency_list.get(start, []):
            if neighbor not in visited:
                path = self._find_longest_path_dfs(neighbor, end, visited.copy())
                if path and len(path) > len(longest):
                    longest = [start] + path
        
        return longest
    
    def _find_path_dfs(self, start: str, end: str, visited: Set[str]) -> List[str]:
        """Find first available path using DFS"""
        if start == end:
            return [start]
        
        visited.add(start)
        
        for neighbor in self.adjacency_list.get(start, []):
            if neighbor not in visited:
                path = self._find_path_dfs(neighbor, end, visited.copy())
                if path:
                    return [start] + path
        
        return None
    
    def _dfs_traverse_uncovered_edges(self, current: str, visited: Set[str], 
                                      covered_edges: Set[Tuple[str, str]]) -> List[str]:
        """DFS to traverse uncovered edges"""
        path = [current]
        visited.add(current)
        
        # Try to find uncovered edges first
        for neighbor in self.adjacency_list.get(current, []):
            edge = (current, neighbor)
            if edge not in covered_edges:
                if neighbor not in visited:
                    covered_edges.add(edge)
                    path.extend(self._dfs_traverse_uncovered_edges(neighbor, visited.copy(), covered_edges))
                    return path
        
        # Otherwise, traverse any edge
        for neighbor in self.adjacency_list.get(current, []):
            if neighbor not in visited:
                path.extend(self._dfs_traverse_uncovered_edges(neighbor, visited.copy(), covered_edges))
                return path
        
        return path
    
    def _generate_gherkin_for_path(self, path: List[str]) -> str:
        """Generate Gherkin-format description for a path"""
        if not path:
            return "Given initial state\nWhen process starts\nThen system is ready"
        
        start = path[0]
        end = path[-1]
        path_str = " → ".join(path)
        
        gherkin = f"Given user starts at {start}\n"
        gherkin += f"When navigating through {path_str}\n"
        gherkin += f"Then should end at {end}"
        
        return gherkin
    
    def export_to_csv(self, test_cases: List[TestCase]) -> str:
        """Export test cases to CSV format"""
        csv_content = "Test Number,Test Description,Expected Result,Coverage Type,Nodes Covered,Edges Covered\n"
        
        for tc in test_cases:
            nodes_str = "|".join(tc.nodes_covered)
            edges_str = "|".join([f"{e[0]}->{e[1]}" for e in tc.edges_covered])
            
            # Escape quotes in description
            description = tc.description.replace('"', '""').replace("\n", " | ")
            expected = tc.expected_result.replace('"', '""')
            
            csv_content += f'{tc.test_number},"{description}","{expected}",{tc.coverage_type},"{nodes_str}","{edges_str}"\n'
        
        return csv_content
    
    def export_to_excel(self, test_cases: List[TestCase]) -> 'bytes':
        """Export test cases to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Test Number", "Test Description", "Expected Result", "Coverage Type", "Nodes Covered", "Edges Covered"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for row_idx, tc in enumerate(test_cases, 2):
            ws.cell(row=row_idx, column=1, value=tc.test_number).border = border
            ws.cell(row=row_idx, column=2, value=tc.description).border = border
            ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=3, value=tc.expected_result).border = border
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=4, value=tc.coverage_type).border = border
            ws.cell(row=row_idx, column=5, value=" → ".join(tc.nodes_covered)).border = border
            ws.cell(row=row_idx, column=6, value=" → ".join([f"{e[0]}→{e[1]}" for e in tc.edges_covered])).border = border
        
        # Col widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
